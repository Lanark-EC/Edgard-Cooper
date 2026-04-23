import re
import pandas as pd
import numpy as np
import io
from datetime import datetime, date
from typing import Optional, List
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


def normalize_col(c) -> str:
    if isinstance(c, (datetime, date)):
        return c.strftime("%d/%m/%Y")
    return str(c).strip()

def is_time_bucket_col(name) -> bool:
    if name is None: return False
    if isinstance(name, (datetime, date)): return True
    s = str(name).strip()
    if not s: return False
    patterns = [
        r"(0?[1-9]|[12]\d|3[01])/(0?[1-9]|1[0-2])/(19|20)\d{2}",
        r"(0?[1-9]|1[0-2])/(19|20)\d{2}",
        r"(19|20)\d{2}",
        r"(19|20)\d{2}[-/](0?[1-9]|1[0-2])",
        r"(19|20)\d{2}[-/](0?[1-9]|1[0-2])[-/](0?[1-9]|[12]\d|3[01])",
        r"(19|20)\d{2}-W\d{1,2}",
        r"(19|20)\d{2}W\d{1,2}",
    ]
    for p in patterns:
        if re.fullmatch(p, s, flags=re.IGNORECASE): return True
    if re.search(r"\bW\d{1,2}\b", s, flags=re.IGNORECASE): return True
    if re.search(r"\b(19|20)\d{2}\b", s) and re.search(r"\bjan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec\b", s.lower()): return True
    return False

def detect_dimensions_and_buckets(df):
    cols = list(df.columns)
    bucket_cols = [c for c in cols if is_time_bucket_col(c)]
    dim_cols    = [c for c in cols if c not in bucket_cols]
    return dim_cols, bucket_cols

def infer_granularity(buckets):
    for b in buckets:
        s = str(b).strip()
        if re.fullmatch(r"(0?[1-9]|[12]\d|3[01])/(0?[1-9]|1[0-2])/(19|20)\d{2}", s): return "weekly"
        if re.search(r"\bW\d{1,2}\b", s, flags=re.IGNORECASE): return "weekly"
    for b in buckets:
        s = str(b).strip()
        if re.fullmatch(r"(0?[1-9]|1[0-2])/(19|20)\d{2}", s): return "monthly"
        if re.fullmatch(r"(19|20)\d{2}[-/](0?[1-9]|1[0-2])", s): return "monthly"
    return "yearly"

def bucket_to_date(bucket) -> Optional[date]:
    if bucket is None: return None
    s = str(bucket).strip()
    if not s: return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try: return datetime.strptime(s, fmt).date()
        except ValueError: pass
    try:
        d = datetime.strptime(s, "%m/%Y")
        return date(d.year, d.month, 1)
    except ValueError: pass
    m = re.fullmatch(r"(19|20)\d{2}[-/](0?[1-9]|1[0-2])", s)
    if m:
        y, mo = re.split(r"[-/]", s)
        return date(int(y), int(mo), 1)
    if re.fullmatch(r"(19|20)\d{2}", s): return date(int(s), 1, 1)
    return None

def pick_sheet_names(file_obj):
    xl = pd.ExcelFile(file_obj, engine="openpyxl")
    sheets = xl.sheet_names
    sales_cands, fcst_cands = [], []
    for sh in sheets:
        low = sh.strip().lower()
        if "actual" in low or "sales" in low: sales_cands.append(sh)
        if "planner forecast" in low:         fcst_cands.append(sh)
    sales_sheet = sales_cands[0] if sales_cands else sheets[0]
    fcst_sheet  = fcst_cands[0]  if fcst_cands  else (sheets[1] if len(sheets)>1 else sheets[0])
    return sales_sheet, fcst_sheet

def run_garvis_export(input_file):
    raw = io.BytesIO(input_file.read())

    sales_sheet, fcst_sheet = pick_sheet_names(raw)
    raw.seek(0)

    sales = pd.read_excel(raw, sheet_name=sales_sheet, engine="openpyxl").dropna(how="all").dropna(axis=1,how="all")
    raw.seek(0)
    fcst  = pd.read_excel(raw, sheet_name=fcst_sheet,  engine="openpyxl").dropna(how="all").dropna(axis=1,how="all")
    raw.seek(0)
    orig_sales_raw = pd.read_excel(raw, sheet_name=sales_sheet, engine="openpyxl")
    raw.seek(0)
    orig_fcst_raw  = pd.read_excel(raw, sheet_name=fcst_sheet,  engine="openpyxl")

    sales.columns = [normalize_col(c) for c in sales.columns]
    fcst.columns  = [normalize_col(c) for c in fcst.columns]

    sales_dims, sales_buckets = detect_dimensions_and_buckets(sales)
    fcst_dims,  fcst_buckets  = detect_dimensions_and_buckets(fcst)

    def non_empty(df, buckets):
        return [b for b in buckets if df[b].replace(0, np.nan).notna().any()]

    sales_buckets_ne = non_empty(sales, sales_buckets)
    fcst_buckets_ne  = non_empty(fcst,  fcst_buckets)

    granularity    = infer_granularity(sales_buckets + fcst_buckets)
    no_sales_window = {"weekly":12,"monthly":3,"yearly":1}.get(granularity,12)

    key_cols = [c for c in sales_dims if c in fcst_dims] or sales_dims[:min(3,len(sales_dims))]

    def make_key(df, cols):
        k = df[cols].copy()
        for c in cols:
            k[c] = k[c].astype(str).str.strip().str.upper()
        return k

    merged = make_key(fcst, key_cols).drop_duplicates()
    merged = merged.merge(make_key(sales, key_cols).drop_duplicates(), on=key_cols, how="outer")

    rows = []
    for _, key_row in merged.iterrows():
        mask_s = pd.Series([True]*len(sales))
        mask_f = pd.Series([True]*len(fcst))
        for c in key_cols:
            mask_s &= (sales[c].astype(str).str.strip().str.upper()==str(key_row[c]))
            mask_f &= (fcst[c].astype(str).str.strip().str.upper() ==str(key_row[c]))
        s_rows = sales[mask_s]
        f_rows = fcst[mask_f]
        row = {c: key_row[c] for c in key_cols}
        for b in sales_buckets_ne:
            row[f"S_{b}"] = float(s_rows[b].sum()) if b in s_rows.columns else 0.0
        for b in fcst_buckets_ne:
            row[f"F_{b}"] = float(f_rows[b].sum()) if b in f_rows.columns else 0.0

        first_fcst_date = next((bucket_to_date(b) for b in fcst_buckets_ne if bucket_to_date(b)), None)
        sales_before = []
        for b in sales_buckets:
            bd = bucket_to_date(b)
            if bd and first_fcst_date and bd < first_fcst_date:
                sales_before.append(float(s_rows[b].sum()) if b in s_rows.columns else 0.0)
        window_vals = sales_before[-no_sales_window:]
        row["12 WKs NO SALES"] = "NO SALES" if window_vals and not any(v>0 for v in window_vals) else ""
        rows.append(row)

    overview = pd.DataFrame(rows)
    s_cols = [c for c in overview.columns if c.startswith("S_")]
    f_cols = [c for c in overview.columns if c.startswith("F_")]

    def safe_avg_from_first(row, cols):
        vals, found = [], False
        for c in cols:
            v = row.get(c, 0)
            if not found and pd.notna(v) and v>0: found = True
            if found and pd.notna(v): vals.append(v)
        return np.mean(vals) if vals else np.nan

    overview["SUM_SALES"]    = overview.apply(lambda r: sum(r[c] for c in s_cols if pd.notna(r.get(c))), axis=1)
    overview["SUM_FORECAST"] = overview.apply(lambda r: sum(r[c] for c in f_cols if pd.notna(r.get(c))), axis=1)
    overview["SUM_DIFF"]     = overview["SUM_SALES"] - overview["SUM_FORECAST"]
    overview["AVG_SALES"]    = overview.apply(lambda r: safe_avg_from_first(r, s_cols), axis=1)
    overview["AVG_FORECAST"] = overview.apply(lambda r: sum(r[c] for c in f_cols if pd.notna(r.get(c)))/len(f_cols) if f_cols else np.nan, axis=1)
    overview["DIFF_FC_SALES"]= overview["AVG_FORECAST"] - overview["AVG_SALES"]
    overview["PCT_DIFF"]     = np.where(overview["AVG_SALES"].fillna(0)!=0, (overview["DIFF_FC_SALES"]/overview["AVG_SALES"])*100, np.nan)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        overview.to_excel(writer, sheet_name="Overview", index=False)
        orig_sales_raw.to_excel(writer, sheet_name="Sales", index=False)
        orig_fcst_raw.to_excel(writer,  sheet_name="Planner Forecast", index=False)

        ws = writer.sheets["Overview"]
        sales_fill = PatternFill("solid", fgColor="FFCCCC")
        fcst_fill  = PatternFill("solid", fgColor="CCE5FF")
        bold       = Font(bold=True)

        for cell in ws[1]:
            cell.font = bold
            h = str(cell.value or "")
            if h.startswith("S_"):
                cell.value = h[2:]
                cell.fill  = sales_fill
            elif h.startswith("F_"):
                cell.value = h[2:]
                cell.fill  = fcst_fill

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len+4, 30)

    buf.seek(0)

    stats = {
        "products": len(overview),
        "sales_periods": len(s_cols),
        "forecast_periods": len(f_cols),
    }
    return buf, stats
