"""
promo_uplift_calc.py — memory-optimized version
Reads only necessary columns, processes in chunks, frees memory aggressively
"""

import pandas as pd
import numpy as np
import io
import re
import gc
from datetime import datetime, timedelta


KNOWN_DIM_COLS = {
    'chain', 'pet', 'subtype', 'productdescription', 'product description',
    'productid', 'product id', 'country', 'ean', 'material', 'sku', 'description'
}

def get_week_cols(df):
    dims = {c for c in df.columns if str(c).strip().lower() in KNOWN_DIM_COLS}
    return [c for c in df.columns
            if c not in dims and re.search(r'\d{2}/\d{2}/\d{4}', str(c).strip())]

def parse_date(col):
    try:
        return datetime.strptime(str(col).strip(), "%d/%m/%Y").date()
    except Exception:
        return None

def find_col(df, candidates):
    col_map = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in col_map:
            return col_map[cand.lower()]
    for cand in candidates:
        for k, v in col_map.items():
            if cand.lower() in k:
                return v
    return None

def read_lean(file_obj, sheet_name, promo_dates=None, extra_dim_cols=None):
    """
    Read Excel keeping only dim cols + relevant week cols.
    promo_dates: set of date objects to keep (None = keep all)
    """
    # Read header only first
    header_df = pd.read_excel(file_obj, sheet_name=sheet_name,
                               nrows=0, engine='openpyxl')
    all_cols = list(header_df.columns)

    # Identify dim cols
    dim_cols = [c for c in all_cols if str(c).strip().lower() in KNOWN_DIM_COLS]
    if extra_dim_cols:
        for ec in extra_dim_cols:
            found = next((c for c in all_cols if str(c).strip().lower() == ec.lower()), None)
            if found and found not in dim_cols:
                dim_cols.append(found)

    # Identify week cols to keep
    week_cols = []
    for c in all_cols:
        if c in dim_cols:
            continue
        d = parse_date(c)
        if d is None:
            continue
        if promo_dates is None or d in promo_dates:
            week_cols.append(c)

    cols_to_read = dim_cols + week_cols

    # Read only needed columns
    if hasattr(file_obj, 'seek'):
        file_obj.seek(0)

    df = pd.read_excel(file_obj, sheet_name=sheet_name,
                       usecols=cols_to_read, engine='openpyxl', dtype=str)
    df.columns = [str(c).strip() for c in df.columns]

    # Convert week cols to numeric
    wc_stripped = [str(c).strip() for c in week_cols]
    for c in wc_stripped:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)

    return df, [str(c).strip() for c in dim_cols], wc_stripped


def run_promo_uplift(ly_file, ty_file, promo_start, promo_end, status_cb=None):
    def status(msg):
        if status_cb:
            status_cb(msg)

    promo_start = promo_start.date() if hasattr(promo_start, 'date') else promo_start
    promo_end   = promo_end.date()   if hasattr(promo_end, 'date')   else promo_end

    # ── Detect sheet names ─────────────────────────────────────
    status("Opening files...")
    ly_xl = pd.ExcelFile(ly_file, engine='openpyxl')
    ty_xl = pd.ExcelFile(ty_file, engine='openpyxl')

    ly_actual_sheet = next((s for s in ly_xl.sheet_names if 'actual' in s.lower()), ly_xl.sheet_names[0])
    ty_fc_sheet     = next((s for s in ty_xl.sheet_names if 'forecast' in s.lower()), None)
    if not ty_fc_sheet:
        raise ValueError("No forecast sheet found in this year's file.")

    # ── Detect TY forecast week cols first (lightweight) ───────
    status("Detecting forecast weeks...")
    if hasattr(ty_file, 'seek'): ty_file.seek(0)
    ty_header = pd.read_excel(ty_file, sheet_name=ty_fc_sheet, nrows=0, engine='openpyxl')
    ty_all_week_cols = [c for c in ty_header.columns
                        if parse_date(c) is not None]
    ty_promo_cols  = [c for c in ty_all_week_cols
                      if parse_date(c) and promo_start <= parse_date(c) <= promo_end]
    ty_normal_cols = [c for c in ty_all_week_cols
                      if parse_date(c) and not (promo_start <= parse_date(c) <= promo_end)]

    if not ty_promo_cols:
        avail = [str(c).strip() for c in ty_all_week_cols[:5]]
        raise ValueError(f"No forecast weeks found between {promo_start} and {promo_end}. Available: {avail}...")

    # ── Detect LY equivalent promo weeks ───────────────────────
    ly_promo_start = promo_start - timedelta(weeks=52)
    ly_promo_end   = promo_end   - timedelta(weeks=52)

    if hasattr(ly_file, 'seek'): ly_file.seek(0)
    ly_header = pd.read_excel(ly_file, sheet_name=ly_actual_sheet, nrows=0, engine='openpyxl')
    ly_all_week_cols = [c for c in ly_header.columns if parse_date(c) is not None]
    ly_promo_cols    = [c for c in ly_all_week_cols
                        if parse_date(c) and ly_promo_start <= parse_date(c) <= ly_promo_end]
    if not ly_promo_cols:
        ly_promo_cols = [c for c in ly_all_week_cols
                         if parse_date(c) and promo_start <= parse_date(c) <= promo_end]
    if not ly_promo_cols:
        raise ValueError(f"No LY weeks found for {ly_promo_start}–{ly_promo_end}.")

    ly_normal_cols = [c for c in ly_all_week_cols
                      if c not in ly_promo_cols]

    status(f"Found {len(ty_promo_cols)} promo weeks ({str(ty_promo_cols[0]).strip()} – {str(ty_promo_cols[-1]).strip()})...")

    # ── Read LY — only promo + normal weeks needed for avg ─────
    # For LY we need ALL weeks to compute the full-year average
    # But we read them all at once since we need avg across all weeks
    status("Reading last year actuals (this takes ~2 min)...")
    if hasattr(ly_file, 'seek'): ly_file.seek(0)
    ly, ly_dim_cols, ly_wc = read_lean(ly_file, ly_actual_sheet)
    gc.collect()

    ly_subtype_col = find_col(ly, ['subtype', 'sub type'])
    ly_country_col = find_col(ly, ['country', 'market'])
    ly_chain_col   = find_col(ly, ['chain', 'customer', 'client'])
    ly_sku_col     = find_col(ly, ['productid', 'product id', 'sku', 'material', 'ean'])
    if not ly_country_col:
        raise ValueError("Country not found in last year file.")
    if not ly_chain_col:
        raise ValueError("Chain not found in last year file.")
    if not ly_sku_col:
        raise ValueError("ProductID/SKU not found in last year file.")

    ly_promo_stripped  = [str(c).strip() for c in ly_promo_cols  if str(c).strip() in ly.columns]
    ly_normal_stripped = [str(c).strip() for c in ly_normal_cols if str(c).strip() in ly.columns]
    ly_all_wc = [c for c in ly.columns if c in set(ly_wc)]

    status("Calculating LY uplift factors...")
    g_ly = [ly_chain_col, ly_sku_col, ly_country_col]

    ly['_avg']   = ly[ly_all_wc].mean(axis=1) if ly_all_wc else 0
    ly['_promo'] = ly[ly_promo_stripped].mean(axis=1) if ly_promo_stripped else 0

    ly_agg = ly.groupby(g_ly, as_index=False).agg(
        ly_avg_weekly=('_avg',   'sum'),
        ly_promo_weekly=('_promo', 'sum')
    )
    ly_agg['ly_uplift'] = np.where(
        ly_agg['ly_avg_weekly'] > 0,
        ly_agg['ly_promo_weekly'] / ly_agg['ly_avg_weekly'],
        np.nan
    )

    del ly; gc.collect()

    # ── Read TY forecast — all weeks for normal avg ─────────── 
    status("Reading this year forecast (this takes ~1 min)...")
    if hasattr(ty_file, 'seek'): ty_file.seek(0)
    ty_fc, ty_dim_cols, ty_wc = read_lean(ty_file, ty_fc_sheet)
    gc.collect()

    ty_subtype_col = find_col(ty_fc, ['subtype', 'sub type'])
    ty_country_col = find_col(ty_fc, ['country', 'market'])
    ty_chain_col   = find_col(ty_fc, ['chain', 'customer', 'client'])
    ty_sku_col     = find_col(ty_fc, ['productid', 'product id', 'sku', 'material', 'ean'])
    if not ty_country_col:
        raise ValueError("Country not found in this year forecast.")
    if not ty_chain_col:
        raise ValueError("Chain not found in this year forecast.")
    if not ty_sku_col:
        raise ValueError("ProductID/SKU not found in this year forecast.")

    ty_promo_stripped  = [str(c).strip() for c in ty_promo_cols  if str(c).strip() in ty_fc.columns]
    ty_normal_stripped = [str(c).strip() for c in ty_normal_cols if str(c).strip() in ty_fc.columns]

    status("Calculating tool uplift factors...")
    g_ty = [ty_chain_col, ty_sku_col, ty_country_col]

    ty_fc['_ty_normal'] = ty_fc[ty_normal_stripped].mean(axis=1) if ty_normal_stripped else 0
    ty_fc['_ty_promo']  = ty_fc[ty_promo_stripped].mean(axis=1)  if ty_promo_stripped  else 0

    ty_agg = ty_fc.groupby(g_ty, as_index=False).agg(
        ty_normal=('_ty_normal', 'sum'),
        ty_promo=('_ty_promo',  'sum')
    )
    ty_agg['tool_uplift'] = np.where(
        ty_agg['ty_normal'] > 0,
        ty_agg['ty_promo'] / ty_agg['ty_normal'],
        np.nan
    )

    # ── Merge and calculate net uplift ─────────────────────────
    status("Calculating net uplift per subtype/country...")
    # Rename before merge to avoid KeyError
    ty_agg = ty_agg.rename(columns={
        ty_chain_col: '_chain', ty_sku_col: '_sku', ty_country_col: '_cty'})
    ly_agg = ly_agg.rename(columns={
        ly_chain_col: '_chain', ly_sku_col: '_sku', ly_country_col: '_cty'})
    merged = ty_agg.merge(ly_agg, on=['_chain', '_sku', '_cty'], how='left')
    merged['net_uplift'] = np.where(
        (merged['tool_uplift'].fillna(0) > 0) & merged['ly_uplift'].notna(),
        merged['ly_uplift'] / merged['tool_uplift'],
        merged['ly_uplift']
    )
    merged['additional_uplift'] = (merged['net_uplift'] - 1.0).clip(lower=0)

    uplift_map = {
        (row['_chain'], row['_sku'], row['_cty']): {
            'ly':   row['ly_uplift'],
            'tool': row['tool_uplift'],
            'add':  row['additional_uplift'],
        }
        for _, row in merged.iterrows()
    }

    # ── Build SKU output ────────────────────────────────────────
    status("Building per-SKU output...")
    OUTPUT_DIMS = ['Chain', 'Pet', 'Subtype', 'ProductDescription', 'ProductID', 'Country']
    ty_output_dims = [(c, find_col(ty_fc, [c])) for c in OUTPUT_DIMS]
    ty_output_dims = [(label, col) for label, col in ty_output_dims if col]

    ty_fc['_sub'] = ty_fc[ty_subtype_col].astype(str).str.strip()
    ty_fc['_cty'] = ty_fc[ty_country_col].astype(str).str.strip()
    promo_col_names = [str(c).strip() for c in ty_promo_cols]

    rows = []
    for _, row in ty_fc.iterrows():
        info = uplift_map.get((str(row.get(ty_chain_col,'')).strip(), str(row.get(ty_sku_col,'')).strip(), str(row.get(ty_country_col,'')).strip()), {})
        add  = info.get('add', np.nan)
        out  = {label: row[col] for label, col in ty_output_dims}
        for wc, wname in zip(ty_promo_stripped, promo_col_names):
            fc_val = float(row[wc]) if pd.notna(row.get(wc)) else 0.0
            out[wname] = round(fc_val * add) if pd.notna(add) and add > 0 else 0
        out['LY uplift']   = round(info['ly'],   3) if pd.notna(info.get('ly'))   else ''
        out['Tool uplift'] = round(info['tool'],  3) if pd.notna(info.get('tool')) else ''
        out['Add. uplift'] = f"{round(add*100,1)}%" if pd.notna(add) else ''
        rows.append(out)

    output_df = pd.DataFrame(rows)

    # Summary
    summary = merged[['_chain','_sku','_cty','ly_avg_weekly','ly_promo_weekly','ly_uplift',
                       'ty_normal','ty_promo','tool_uplift','additional_uplift']].copy()
    summary.columns = ['Chain','SKU','Country','LY avg weekly','LY promo weekly','LY uplift factor',
                        'TY avg weekly forecast','TY promo weekly forecast','Tool uplift factor',
                        'Additional uplift']
    summary['Additional uplift'] = summary['Additional uplift'].apply(
        lambda x: f"{round(x*100,1)}%" if pd.notna(x) else '')

    del ty_fc, merged; gc.collect()

    # ── Write Excel ─────────────────────────────────────────────
    status("Writing Excel output...")
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        output_df.to_excel(writer, sheet_name='Promo Registration', index=False)
        summary.to_excel(writer, sheet_name='Uplift Summary', index=False)

        from openpyxl.styles import PatternFill, Font
        ws = writer.sheets['Promo Registration']
        teal_fill  = PatternFill('solid', fgColor='3BBFBF')
        week_fill  = PatternFill('solid', fgColor='E3F7F7')
        white_bold = Font(bold=True, color='FFFFFF')
        for cell in ws[1]:
            col_name = str(cell.value or '')
            if re.search(r'\d{2}/\d{2}/\d{4}', col_name):
                cell.fill = week_fill
                cell.font = Font(bold=True)
            else:
                cell.fill = teal_fill
                cell.font = white_bold
        for col in ws.columns:
            ml = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(ml + 3, 25)

    buf.seek(0)

    stats = {
        'promo_weeks':   len(ty_promo_stripped),
        'sku_count':     len(output_df),
        'skus':          int(summary['SKU'].nunique()),
        'countries':     int(summary['Country'].nunique()),
        'avg_ly_uplift': round(float(summary['LY uplift factor'].replace('',np.nan).dropna().astype(float).mean()), 2)
                         if summary['LY uplift factor'].replace('',np.nan).dropna().any() else None,
    }
    return buf, stats
